import pandas as pd
import common as c
import datetime as dt
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def generate_schedule(file_name):
    amount          =    c.Get_amount()
    life            =    c.get_life()
    entry_date      =    c.Entery()
    retiring_date   =    c.Retiring(entry_date)
    life_cycle      =    c.life_cycle(entry_date,retiring_date,life)
    constant_rate   =    c.Constant(life)

    #creating a date series monthly
    Monthly_dates = pd.Series(pd.date_range(entry_date,freq="ME",periods=life_cycle )).dt.strftime('%Y-%m-%d')

    #calculating monthly basis
    Monthly_records = []
    depriciation_value= c.monthly_depreciation(amount,constant_rate)
    accumulated = 0
    condition=False
    for date in Monthly_dates :
        accumulated+= depriciation_value
        Vna= amount - accumulated
        if Vna <= 1 :
            Vna = 0
            condition = True
        Monthly_records .append({
            "Date": date,
            "Amount": amount,
            "Rate": constant_rate,
            "Depreciation": depriciation_value,
            "Accumulation": accumulated,
            "VNA": Vna
        })
        if condition == True :
            break
    monthlytable    = pd.DataFrame(Monthly_records)


    #calculating Yearly basis
    Yearly_records = monthlytable.copy()
    Yearly_records["Date"] = pd.to_datetime(Yearly_records["Date"])
    Yearly_records["Year"] = Yearly_records["Date"].dt.year
    Yearly_records = Yearly_records.groupby("Year").aggregate({
        "Date"          : "last",
        "Amount"        : "last",
        "Rate"          : "last",
        "Depreciation"  : "sum",
        "Accumulation"  : "last",
        "VNA"           : "last"
    })
    Yearly_records.iat[0, Yearly_records.columns.get_loc("Date")] = entry_date

    #exporting both outputs

    yearlytable     = pd.DataFrame(Yearly_records)

    # ==============================================================================
    # EXPORT & STYLING BLOCK
    # ==============================================================================
    file_path = "depreciation_table.xlsx"

    file_path = file_name

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        yearlytable.to_excel(writer, sheet_name="Yearly", index=False)
        monthlytable.to_excel(writer, sheet_name="Monthly", index=False)

        wb = writer.book

        # Number format without currency symbol
        accounting_fmt = "#,##0.00"

        # Header styling (Dark Navy Blue)
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )

        data_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)

        # Dark borders for clear cell separation
        cell_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        header_border = Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            top=Side(style="medium", color="1F4E79"),
            bottom=Side(style="medium", color="000000"),
        )

        double_bottom_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="double", color="000000"),
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # Soft Pastel Color Palette for Years
        year_colors = [
            "F2F2F2",
            "E2EFDA",
            "DDEBF7",
            "FFF2CC",
            "FCE4D6",
            "E8D8F8",
            "E0F2F1",
            "FFE0B2",
        ]

        # Map each year to a specific color
        unique_years = (
            pd.to_datetime(yearlytable["Date"]).dt.year.unique().tolist()
        )
        year_color_map = {
            year: year_colors[i % len(year_colors)]
            for i, year in enumerate(unique_years)
        }

        # Style Yearly and Monthly sheets
        for sheet_name in ["Yearly", "Monthly"]:
            ws = writer.sheets[sheet_name]
            ws.views.sheetView[0].showGridLines = True

            # Style Headers
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = header_border

            # Style Rows
            for row in range(2, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                row_year = pd.to_datetime(date_val).year
                row_fill = PatternFill(
                    start_color=year_color_map[row_year],
                    end_color=year_color_map[row_year],
                    fill_type="solid",
                )

                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.border = cell_border

                    col_name = str(ws.cell(row=1, column=col).value)

                    if "Date" in col_name:
                        cell.number_format = "YYYY-MM-DD"
                        cell.alignment = align_center
                    elif "Rate" in col_name:
                        cell.number_format = "0.00%"
                        cell.alignment = align_right
                    else:
                        cell.number_format = accounting_fmt
                        cell.alignment = align_right

            # Final Row Double Underline
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = bold_font
                cell.border = double_bottom_border

            # Auto-fit Columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 14)