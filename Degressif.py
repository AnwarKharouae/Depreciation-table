import  common as c
import pandas as pd
import datetime as dt
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
def generate_schedule(file_name):
    amount              =    c.Get_amount()
    #making sure  its not less than 3 year
    while True :
        try :
            life              =    c.get_life()
            constant_rate     =    c.Constant(life)
            constant_rate     =    c.rate(life,constant_rate)
            break
        except ValueError:
            print("PLease Note that the declining rate asset life should not be less than 3 years")

    entry_date          =    c.Entery()
    retiring_date       =    c.Retiring(entry_date)
    life_cycle          =    c.life_cycle(entry_date,retiring_date,life)




    #creating a date series monthly/yearly
    Monthly_dates   = pd.Series(pd.date_range(entry_date,freq="ME",periods=life_cycle )).dt.strftime('%Y-%m-%d')

    Yearly_dates    = Monthly_dates.copy()
    Yearly_dates    = pd.to_datetime(Yearly_dates)
    Yearly_records  = Yearly_dates.dt.year 

    #creating a combined df for later use
    dates           = pd.DataFrame({
        "Date"      : Yearly_records,
        "Monthly"   : Monthly_dates
    })

    #having a date for each year
    dates_Y     = dates.groupby("Date").aggregate({

        "Monthly"   : "last"
    }).reset_index()

    #identifying the number of months to help calculating
    months          = []
    months          = Yearly_records.value_counts().sort_index()

    #variable declaration
    accumulated           =0
    condition_rate        = False
    declining_rate        =c.Constant(life)
    i                     = 0
    records               =[]
    condition             =False
    Final_Yearly_record   =[]

    #filling the first column in records
    Yearly_dates_values = dates_Y["Monthly"].tolist()
    Yearly_dates_values[0] = entry_date.strftime("%Y-%m-%d")
    X_amount = amount
    #calculating yearly basis
    for month,date in zip(months,Yearly_dates_values):
        
        #rate cheking
        if constant_rate > declining_rate :
            rate=constant_rate
        else :
            rate=declining_rate

        # Calculating depreciation       
        depriciation_value  =   c.Yearly_depreciation(X_amount,rate,month)
        accumulated         +=  depriciation_value
        Vna                 =  X_amount - depriciation_value

        #checking last month so we dont have a negative Vna
        if Vna <= 1 :
            Vna = 0
            condition = True

        # main loop
        Final_Yearly_record.append({
            "Date"          : date,
            "Amount"        : X_amount,
            "Rate"          : rate,
            "Depreciation"  : depriciation_value,
            "Accumulation"  : accumulated,
            "VNA"           : Vna
        })

        #if vna is less than 1 it will break the loop
        if condition == True :
            break

        X_amount = Vna

        #calculating declining rate
        if condition_rate == True :
            break
        # just me checking

        x=life-month
        print(life,month)
        declining_rate      = 12/x
        life                -=month
        print(declining_rate)

        if declining_rate >= 1 :
            condition_rate = True



    #calculating monthly basis
    final_monthly_record= []
    Purchase_value      = amount 
    accumulated         = 0
    X_condition         = False
    for id, row in dates.iterrows():

        #getting year from current row
        current_year = row["Date"]

        #getting values from final yearly records
        yearly_item = next(
            item for item in Final_Yearly_record 
            if pd.to_datetime(item["Date"]).year == current_year
        )
        rate = yearly_item["Rate"]
        amount = yearly_item["Amount"]

        # monthly depreciation 
        depreciation = rate*amount*1/12
        accumulated += depreciation
        Vna          = Purchase_value-accumulated
        if Vna <= 1 :
            Vna = 0
            X_condition = True

        #adding row to table 
        final_monthly_record.append({
            "Date": row["Monthly"],
            "Amount": amount,
            "Rate": rate,
            "Depreciation": depreciation,
            "Accumulation": accumulated ,
            "VNA": Vna
        })
        if X_condition == True :
            break

    # Turn into DataFrame
    df_monthly = pd.DataFrame(final_monthly_record)

    Final_Yearly_record= pd.DataFrame(Final_Yearly_record)




    #exporting
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        Final_Yearly_record.to_excel(writer, sheet_name="Yearly", index=False)
        df_monthly.to_excel(writer, sheet_name="Monthly", index=False)



    # ==============================================================================
    # EXPORT & STYLING BLOCK (thanks to the AI)
    # ==============================================================================
    file_path = file_name

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        Final_Yearly_record.to_excel(writer, sheet_name="Yearly", index=False)
        df_monthly.to_excel(writer, sheet_name="Monthly", index=False)

        wb = writer.book

        # Number format without currency symbol
        accounting_fmt = "#,##0.00"

        # Header styling (Dark Navy Blue)
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )

        data_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)

        # Dark borders for clear cell separation
        cell_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        header_border = Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            top=Side(style="medium", color="1F4E79"),
            bottom=Side(style="medium", color="000000"),
        )

        double_bottom_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="double", color="000000"),
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # Soft Pastel Color Palette for Years
        year_colors = [
            "F2F2F2",
            "E2EFDA",
            "DDEBF7",
            "FFF2CC",
            "FCE4D6",
            "E8D8F8",
            "E0F2F1",
            "FFE0B2",
        ]

        # Map each year to a specific color
        unique_years = (
            pd.to_datetime(Final_Yearly_record["Date"]).dt.year.unique().tolist()
        )
        year_color_map = {
            year: year_colors[i % len(year_colors)]
            for i, year in enumerate(unique_years)
        }

        # Style Yearly and Monthly sheets
        for sheet_name in ["Yearly", "Monthly"]:
            ws = writer.sheets[sheet_name]
            ws.views.sheetView[0].showGridLines = True

            # Style Headers
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = header_border

            # Style Rows
            for row in range(2, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                row_year = pd.to_datetime(date_val).year
                row_fill = PatternFill(
                    start_color=year_color_map[row_year],
                    end_color=year_color_map[row_year],
                    fill_type="solid",
                )

                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.border = cell_border

                    col_name = str(ws.cell(row=1, column=col).value)

                    if "Date" in col_name:
                        cell.number_format = "YYYY-MM-DD"
                        cell.alignment = align_center
                    elif "Rate" in col_name:
                        cell.number_format = "0.00%"
                        cell.alignment = align_right
                    else:
                        cell.number_format = accounting_fmt
                        cell.alignment = align_right

            # Final Row Double Underline
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = bold_font
                cell.border = double_bottom_border

            # Auto-fit Columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 14)